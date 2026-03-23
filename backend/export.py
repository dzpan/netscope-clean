"""Export functions: DrawIO XML, CSV zip, Excel, DOT, SVG, JSON."""

from __future__ import annotations

import csv
import io
import math
import shutil
import subprocess
import zipfile
from datetime import UTC
from typing import TypeAlias
from xml.sax.saxutils import escape

import openpyxl
import openpyxl.worksheet.worksheet
from openpyxl.styles import Alignment, Font, PatternFill

from backend.models import TopologyResult

_WS: TypeAlias = openpyxl.worksheet.worksheet.Worksheet

# ---------------------------------------------------------------------------
# DrawIO
# ---------------------------------------------------------------------------

_DRAWIO_HEADER = """\
<?xml version="1.0" encoding="UTF-8"?>
<mxfile host="NetScope" modified="{modified}" type="device">
  <diagram id="topology" name="Network Topology">
    <mxGraphModel dx="1422" dy="762" grid="1" gridSize="10" guides="1" tooltips="1"
      connect="1" arrows="1" fold="1" page="1" pageScale="1"
      pageWidth="1169" pageHeight="827" math="0" shadow="0">
      <root>
        <mxCell id="0"/>
        <mxCell id="1" parent="0"/>
"""
_DRAWIO_FOOTER = """\
      </root>
    </mxGraphModel>
  </diagram>
</mxfile>
"""

_DEVICE_COLORS = {
    "ok": "dae8fc",
    "placeholder": "f5f5f5",
    "unreachable": "f8cecc",
    "auth_failed": "ffe6cc",
    "timeout": "ffe6cc",
    "no_cdp_lldp": "fff2cc",
}

_DEVICE_STROKE = {
    "ok": "6c8ebf",
    "placeholder": "666666",
    "unreachable": "b85450",
    "auth_failed": "d6b656",
    "timeout": "d6b656",
    "no_cdp_lldp": "d6b656",
}


def _abbreviate_intf(name: str) -> str:
    """Shorten interface names for diagram readability."""
    for long, short in (
        ("TenGigabitEthernet", "Te"),
        ("GigabitEthernet", "Gi"),
        ("FastEthernet", "Fa"),
        ("Port-channel", "Po"),
        ("Ethernet", "Eth"),
    ):
        if name.startswith(long):
            return short + name[len(long) :]
    return name


def _make_cell_id(device_id: str) -> str:
    """Sanitize device ID into a valid XML id attribute."""
    import re

    return "dev_" + re.sub(r"[^A-Za-z0-9_]", "_", device_id)


def export_drawio(result: TopologyResult) -> str:
    from datetime import datetime

    header = _DRAWIO_HEADER.format(
        modified=datetime.now(UTC).strftime("%Y-%m-%dT%H:%M:%S.000Z"),
    )
    cells: list[str] = [header]
    n = len(result.devices)
    indent = " " * 8  # inside <root>

    # Simple circular layout
    for i, device in enumerate(result.devices):
        angle = 2 * math.pi * i / max(n, 1)
        radius = max(200, 80 * n // 6)
        cx = 600 + radius * math.cos(angle)
        cy = 400 + radius * math.sin(angle)
        x, y = cx - 60, cy - 30

        status = device.status.value
        fill = _DEVICE_COLORS.get(status, "dae8fc")
        stroke = _DEVICE_STROKE.get(status, "6c8ebf")
        label = escape(device.hostname or device.id)
        sub = escape(device.mgmt_ip)

        cell_id = _make_cell_id(device.id)
        cells.append(
            f'{indent}<mxCell id="{cell_id}" value="&lt;b&gt;{label}&lt;/b&gt;&lt;br/&gt;'
            f'&lt;font style=&quot;font-size:10px&quot;&gt;{sub}&lt;/font&gt;" '
            f'style="rounded=1;whiteSpace=wrap;html=1;fillColor=#{fill};strokeColor=#{stroke};'
            f'fontSize=12;" vertex="1" parent="1">\n'
            f'{indent}  <mxGeometry x="{x:.0f}" y="{y:.0f}"'
            f' width="140" height="50" as="geometry"/>\n'
            f"{indent}</mxCell>\n"
        )

    for j, link in enumerate(result.links):
        src_id = _make_cell_id(link.source)
        tgt_id = _make_cell_id(link.target)
        src_intf = _abbreviate_intf(link.source_intf)
        tgt_intf = _abbreviate_intf(link.target_intf or "")
        label = escape(f"{src_intf} - {tgt_intf}")
        cells.append(
            f'{indent}<mxCell id="link_{j}" value="{label}" '
            f'style="edgeStyle=orthogonalEdgeStyle;fontSize=9;fontColor=#999999;" '
            f'edge="1" source="{src_id}" target="{tgt_id}" parent="1">\n'
            f'{indent}  <mxGeometry relative="1" as="geometry"/>\n'
            f"{indent}</mxCell>\n"
        )

    cells.append(_DRAWIO_FOOTER)
    return "".join(cells)


# ---------------------------------------------------------------------------
# CSV zip
# ---------------------------------------------------------------------------


def export_csv_zip(result: TopologyResult) -> bytes:
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("devices.csv", _devices_csv(result))
        zf.writestr("links.csv", _links_csv(result))
        zf.writestr("failures.csv", _failures_csv(result))
        zf.writestr("interfaces.csv", _interfaces_csv(result))
        zf.writestr("vlans.csv", _vlans_csv(result))
        zf.writestr("arp_table.csv", _arp_table_csv(result))
        zf.writestr("mac_table.csv", _mac_table_csv(result))
        zf.writestr("routes.csv", _routes_csv(result))
        zf.writestr("trunks.csv", _trunks_csv(result))
        zf.writestr("etherchannels.csv", _etherchannels_csv(result))
        zf.writestr("stp_vlans.csv", _stp_vlans_csv(result))
        zf.writestr("stp_ports.csv", _stp_ports_csv(result))
        zf.writestr("nve_peers.csv", _nve_peers_csv(result))
        zf.writestr("vni_mappings.csv", _vni_mappings_csv(result))
        zf.writestr("evpn_neighbors.csv", _evpn_neighbors_csv(result))
    return buf.getvalue()


def _devices_csv(result: TopologyResult) -> str:
    out = io.StringIO()
    w = csv.writer(out)
    w.writerow(["ID", "Hostname", "Management IP", "Platform", "Serial", "OS Version", "Status"])
    for d in result.devices:
        w.writerow(
            [
                d.id,
                d.hostname,
                d.mgmt_ip,
                d.platform,
                d.serial,
                d.os_version,
                d.status.value,
            ]
        )
    return out.getvalue()


def _links_csv(result: TopologyResult) -> str:
    out = io.StringIO()
    w = csv.writer(out)
    w.writerow(["Source", "Source Interface", "Target", "Target Interface", "Protocol"])
    for lk in result.links:
        w.writerow([lk.source, lk.source_intf, lk.target, lk.target_intf, lk.protocol])
    return out.getvalue()


def _failures_csv(result: TopologyResult) -> str:
    out = io.StringIO()
    w = csv.writer(out)
    w.writerow(["Target", "Reason", "Detail"])
    for f in result.failures:
        w.writerow([f.target, f.reason, f.detail])
    return out.getvalue()


def _interfaces_csv(result: TopologyResult) -> str:
    out = io.StringIO()
    w = csv.writer(out)
    w.writerow(
        ["Device", "Interface", "Status", "VLAN", "Speed", "Duplex", "Description", "IP Address"]
    )
    for d in result.devices:
        for intf in d.interfaces:
            w.writerow(
                [
                    d.id,
                    intf.name,
                    intf.status,
                    intf.vlan,
                    intf.speed,
                    intf.duplex,
                    intf.description,
                    intf.ip_address,
                ]
            )
    return out.getvalue()


def _vlans_csv(result: TopologyResult) -> str:
    out = io.StringIO()
    w = csv.writer(out)
    w.writerow(["Device", "VLAN ID", "Name", "Status"])
    for d in result.devices:
        for vl in d.vlans:
            w.writerow([d.id, vl.vlan_id, vl.name, vl.status])
    return out.getvalue()


def _arp_table_csv(result: TopologyResult) -> str:
    out = io.StringIO()
    w = csv.writer(out)
    w.writerow(["Device", "IP Address", "MAC Address", "Interface", "Type"])
    for d in result.devices:
        for arp in d.arp_table:
            w.writerow([d.id, arp.ip_address, arp.mac_address, arp.interface, arp.entry_type])
    return out.getvalue()


def _mac_table_csv(result: TopologyResult) -> str:
    out = io.StringIO()
    w = csv.writer(out)
    w.writerow(["Device", "MAC Address", "VLAN", "Interface", "Type"])
    for d in result.devices:
        for mac in d.mac_table:
            w.writerow([d.id, mac.mac_address, mac.vlan_id, mac.interface, mac.entry_type])
    return out.getvalue()


def _routes_csv(result: TopologyResult) -> str:
    out = io.StringIO()
    w = csv.writer(out)
    w.writerow(
        ["Device", "Protocol", "Route Type", "Destination", "Next Hop", "Interface", "Metric"]
    )
    for d in result.devices:
        for rt in d.route_table:
            w.writerow(
                [
                    d.id,
                    rt.protocol,
                    rt.route_type,
                    rt.destination,
                    rt.next_hop,
                    rt.interface,
                    rt.metric,
                ]
            )
    return out.getvalue()


def _trunks_csv(result: TopologyResult) -> str:
    out = io.StringIO()
    w = csv.writer(out)
    w.writerow(
        [
            "Device",
            "Port",
            "Mode",
            "Encapsulation",
            "Status",
            "Native VLAN",
            "Allowed VLANs",
            "Active VLANs",
            "Forwarding VLANs",
        ]
    )
    for d in result.devices:
        for port_name, tr in d.trunk_info.items():
            w.writerow(
                [
                    d.id,
                    port_name,
                    tr.mode,
                    tr.encapsulation,
                    tr.status,
                    tr.native_vlan,
                    tr.allowed_vlans,
                    tr.active_vlans,
                    tr.forwarding_vlans,
                ]
            )
    return out.getvalue()


def _etherchannels_csv(result: TopologyResult) -> str:
    out = io.StringIO()
    w = csv.writer(out)
    w.writerow(
        [
            "Device",
            "Channel ID",
            "Port-Channel",
            "Status",
            "Protocol",
            "Layer",
            "Member Port",
            "Member Flag",
            "Member Status",
        ]
    )
    for d in result.devices:
        for ec in d.etherchannels:
            if ec.members:
                for member in ec.members:
                    w.writerow(
                        [
                            d.id,
                            ec.channel_id,
                            ec.port_channel,
                            ec.status,
                            ec.protocol,
                            ec.layer,
                            member.interface,
                            member.status,
                            member.status_desc,
                        ]
                    )
            else:
                w.writerow(
                    [
                        d.id,
                        ec.channel_id,
                        ec.port_channel,
                        ec.status,
                        ec.protocol,
                        ec.layer,
                        "",
                        "",
                        "",
                    ]
                )
    return out.getvalue()


def _stp_vlans_csv(result: TopologyResult) -> str:
    out = io.StringIO()
    w = csv.writer(out)
    w.writerow(
        [
            "Device",
            "VLAN",
            "Protocol",
            "Is Root",
            "Root Priority",
            "Root Address",
            "Root Cost",
            "Bridge Priority",
            "Bridge Address",
        ]
    )
    for d in result.devices:
        for sv in d.stp_info:
            w.writerow(
                [
                    d.id,
                    sv.vlan_id,
                    sv.protocol,
                    sv.is_root,
                    sv.root_priority,
                    sv.root_address,
                    sv.root_cost,
                    sv.bridge_priority,
                    sv.bridge_address,
                ]
            )
    return out.getvalue()


def _stp_ports_csv(result: TopologyResult) -> str:
    out = io.StringIO()
    w = csv.writer(out)
    w.writerow(["Device", "VLAN", "Interface", "Role", "State", "Cost", "Priority", "Type"])
    for d in result.devices:
        for sv in d.stp_info:
            for p in sv.ports:
                w.writerow(
                    [
                        d.id,
                        sv.vlan_id,
                        p.interface,
                        p.role,
                        p.state,
                        p.cost,
                        p.port_priority,
                        p.link_type,
                    ]
                )
    return out.getvalue()


def _nve_peers_csv(result: TopologyResult) -> str:
    out = io.StringIO()
    w = csv.writer(out)
    w.writerow(["Device", "Interface", "Peer IP", "State", "Learn Type", "Uptime", "Router MAC"])
    for d in result.devices:
        for p in d.nve_peers:
            w.writerow(
                [
                    d.id,
                    p.interface,
                    p.peer_ip,
                    p.state,
                    p.learn_type,
                    p.uptime,
                    p.router_mac,
                ]
            )
    return out.getvalue()


def _vni_mappings_csv(result: TopologyResult) -> str:
    out = io.StringIO()
    w = csv.writer(out)
    w.writerow(["Device", "Interface", "VNI", "Multicast Group", "State", "Mode", "Type", "BD/VRF"])
    for d in result.devices:
        for v in d.vni_mappings:
            w.writerow(
                [
                    d.id,
                    v.interface,
                    v.vni,
                    v.multicast_group,
                    v.state,
                    v.mode,
                    v.vni_type,
                    v.bd_vrf,
                ]
            )
    return out.getvalue()


def _evpn_neighbors_csv(result: TopologyResult) -> str:
    out = io.StringIO()
    w = csv.writer(out)
    w.writerow(
        [
            "Device",
            "Neighbor",
            "ASN",
            "Version",
            "Msg Rcvd",
            "Msg Sent",
            "Up/Down",
            "State/PfxRcv",
        ]
    )
    for d in result.devices:
        for e in d.evpn_neighbors:
            w.writerow(
                [
                    d.id,
                    e.neighbor,
                    e.asn,
                    e.version,
                    e.msg_rcvd,
                    e.msg_sent,
                    e.up_down,
                    e.state_pfx_rcv,
                ]
            )
    return out.getvalue()


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------

_HEADER_FILL = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True)


def export_excel(result: TopologyResult) -> bytes:
    wb = openpyxl.Workbook()

    _write_devices_sheet(wb.active, result)
    wb.active.title = "Devices"

    _write_links_sheet(wb.create_sheet("Links"), result)
    _write_failures_sheet(wb.create_sheet("Failures"), result)
    _write_interfaces_sheet(wb.create_sheet("Interfaces"), result)
    _write_vlans_sheet(wb.create_sheet("VLANs"), result)
    _write_arp_table_sheet(wb.create_sheet("ARP Table"), result)
    _write_mac_table_sheet(wb.create_sheet("MAC Table"), result)
    _write_routes_sheet(wb.create_sheet("Routes"), result)
    _write_trunks_sheet(wb.create_sheet("Trunks"), result)
    _write_etherchannels_sheet(wb.create_sheet("EtherChannels"), result)
    _write_stp_vlans_sheet(wb.create_sheet("STP VLANs"), result)
    _write_stp_ports_sheet(wb.create_sheet("STP Ports"), result)
    _write_nve_peers_sheet(wb.create_sheet("NVE Peers"), result)
    _write_vni_mappings_sheet(wb.create_sheet("VNI Mappings"), result)
    _write_evpn_neighbors_sheet(wb.create_sheet("EVPN Neighbors"), result)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_devices_sheet(ws: _WS, result: TopologyResult) -> None:
    headers = ["ID", "Hostname", "Management IP", "Platform", "Serial", "OS Version", "Status"]
    _write_header_row(ws, headers)
    for d in result.devices:
        ws.append([d.id, d.hostname, d.mgmt_ip, d.platform, d.serial, d.os_version, d.status.value])
    _auto_width(ws)


def _write_links_sheet(ws: _WS, result: TopologyResult) -> None:
    headers = ["Source", "Source Interface", "Target", "Target Interface", "Protocol"]
    _write_header_row(ws, headers)
    for lk in result.links:
        ws.append([lk.source, lk.source_intf, lk.target, lk.target_intf, lk.protocol])
    _auto_width(ws)


def _write_failures_sheet(ws: _WS, result: TopologyResult) -> None:
    headers = ["Target", "Reason", "Detail"]
    _write_header_row(ws, headers)
    for f in result.failures:
        ws.append([f.target, f.reason, f.detail])
    _auto_width(ws)


def _write_interfaces_sheet(ws: _WS, result: TopologyResult) -> None:
    headers = [
        "Device",
        "Interface",
        "Status",
        "VLAN",
        "Speed",
        "Duplex",
        "Description",
        "IP Address",
    ]
    _write_header_row(ws, headers)
    for d in result.devices:
        for intf in d.interfaces:
            ws.append(
                [
                    d.id,
                    intf.name,
                    intf.status,
                    intf.vlan,
                    intf.speed,
                    intf.duplex,
                    intf.description,
                    intf.ip_address,
                ]
            )
    _auto_width(ws)


def _write_vlans_sheet(ws: _WS, result: TopologyResult) -> None:
    headers = ["Device", "VLAN ID", "Name", "Status"]
    _write_header_row(ws, headers)
    for d in result.devices:
        for vl in d.vlans:
            ws.append([d.id, vl.vlan_id, vl.name, vl.status])
    _auto_width(ws)


def _write_arp_table_sheet(ws: _WS, result: TopologyResult) -> None:
    headers = ["Device", "IP Address", "MAC Address", "Interface", "Type"]
    _write_header_row(ws, headers)
    for d in result.devices:
        for arp in d.arp_table:
            ws.append([d.id, arp.ip_address, arp.mac_address, arp.interface, arp.entry_type])
    _auto_width(ws)


def _write_mac_table_sheet(ws: _WS, result: TopologyResult) -> None:
    headers = ["Device", "MAC Address", "VLAN", "Interface", "Type"]
    _write_header_row(ws, headers)
    for d in result.devices:
        for mac in d.mac_table:
            ws.append([d.id, mac.mac_address, mac.vlan_id, mac.interface, mac.entry_type])
    _auto_width(ws)


def _write_routes_sheet(ws: _WS, result: TopologyResult) -> None:
    headers = ["Device", "Protocol", "Route Type", "Destination", "Next Hop", "Interface", "Metric"]
    _write_header_row(ws, headers)
    for d in result.devices:
        for rt in d.route_table:
            ws.append(
                [
                    d.id,
                    rt.protocol,
                    rt.route_type,
                    rt.destination,
                    rt.next_hop,
                    rt.interface,
                    rt.metric,
                ]
            )
    _auto_width(ws)


def _write_trunks_sheet(ws: _WS, result: TopologyResult) -> None:
    headers = [
        "Device",
        "Port",
        "Mode",
        "Encapsulation",
        "Status",
        "Native VLAN",
        "Allowed VLANs",
        "Active VLANs",
        "Forwarding VLANs",
    ]
    _write_header_row(ws, headers)
    for d in result.devices:
        for port_name, tr in d.trunk_info.items():
            ws.append(
                [
                    d.id,
                    port_name,
                    tr.mode,
                    tr.encapsulation,
                    tr.status,
                    tr.native_vlan,
                    tr.allowed_vlans,
                    tr.active_vlans,
                    tr.forwarding_vlans,
                ]
            )
    _auto_width(ws)


def _write_etherchannels_sheet(ws: _WS, result: TopologyResult) -> None:
    headers = [
        "Device",
        "Channel ID",
        "Port-Channel",
        "Status",
        "Protocol",
        "Layer",
        "Member Port",
        "Member Flag",
        "Member Status",
    ]
    _write_header_row(ws, headers)
    for d in result.devices:
        for ec in d.etherchannels:
            if ec.members:
                for member in ec.members:
                    ws.append(
                        [
                            d.id,
                            ec.channel_id,
                            ec.port_channel,
                            ec.status,
                            ec.protocol,
                            ec.layer,
                            member.interface,
                            member.status,
                            member.status_desc,
                        ]
                    )
            else:
                ws.append(
                    [
                        d.id,
                        ec.channel_id,
                        ec.port_channel,
                        ec.status,
                        ec.protocol,
                        ec.layer,
                        "",
                        "",
                        "",
                    ]
                )
    _auto_width(ws)


def _write_stp_vlans_sheet(ws: _WS, result: TopologyResult) -> None:
    headers = [
        "Device",
        "VLAN",
        "Protocol",
        "Is Root",
        "Root Priority",
        "Root Address",
        "Root Cost",
        "Bridge Priority",
        "Bridge Address",
    ]
    _write_header_row(ws, headers)
    for d in result.devices:
        for sv in d.stp_info:
            ws.append(
                [
                    d.id,
                    sv.vlan_id,
                    sv.protocol,
                    sv.is_root,
                    sv.root_priority,
                    sv.root_address,
                    sv.root_cost,
                    sv.bridge_priority,
                    sv.bridge_address,
                ]
            )
    _auto_width(ws)


def _write_stp_ports_sheet(ws: _WS, result: TopologyResult) -> None:
    headers = ["Device", "VLAN", "Interface", "Role", "State", "Cost", "Priority", "Type"]
    _write_header_row(ws, headers)
    for d in result.devices:
        for sv in d.stp_info:
            for p in sv.ports:
                ws.append(
                    [
                        d.id,
                        sv.vlan_id,
                        p.interface,
                        p.role,
                        p.state,
                        p.cost,
                        p.port_priority,
                        p.link_type,
                    ]
                )
    _auto_width(ws)


def _write_nve_peers_sheet(ws: _WS, result: TopologyResult) -> None:
    headers = ["Device", "Interface", "Peer IP", "State", "Learn Type", "Uptime", "Router MAC"]
    _write_header_row(ws, headers)
    for d in result.devices:
        for p in d.nve_peers:
            ws.append([d.id, p.interface, p.peer_ip, p.state, p.learn_type, p.uptime, p.router_mac])
    _auto_width(ws)


def _write_vni_mappings_sheet(ws: _WS, result: TopologyResult) -> None:
    headers = ["Device", "Interface", "VNI", "Multicast Group", "State", "Mode", "Type", "BD/VRF"]
    _write_header_row(ws, headers)
    for d in result.devices:
        for v in d.vni_mappings:
            ws.append(
                [
                    d.id,
                    v.interface,
                    v.vni,
                    v.multicast_group,
                    v.state,
                    v.mode,
                    v.vni_type,
                    v.bd_vrf,
                ]
            )
    _auto_width(ws)


def _write_evpn_neighbors_sheet(ws: _WS, result: TopologyResult) -> None:
    headers = [
        "Device",
        "Neighbor",
        "ASN",
        "Version",
        "Msg Rcvd",
        "Msg Sent",
        "Up/Down",
        "State/PfxRcv",
    ]
    _write_header_row(ws, headers)
    for d in result.devices:
        for e in d.evpn_neighbors:
            ws.append(
                [
                    d.id,
                    e.neighbor,
                    e.asn,
                    e.version,
                    e.msg_rcvd,
                    e.msg_sent,
                    e.up_down,
                    e.state_pfx_rcv,
                ]
            )
    _auto_width(ws)


def _write_header_row(ws: _WS, headers: list[str]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def _auto_width(ws: _WS) -> None:
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)


# ---------------------------------------------------------------------------
# DOT (Graphviz)
# ---------------------------------------------------------------------------


def export_dot(result: TopologyResult) -> str:
    """Generate a Graphviz DOT undirected graph from topology data."""
    lines = ["graph topology {", "  overlap=false;", "  splines=true;"]
    for device in result.devices:
        label = f"{device.hostname or device.id}\\n{device.mgmt_ip}"
        lines.append(f'  "{device.id}" [label="{label}"];')
    seen: set[frozenset[str]] = set()
    for link in result.links:
        key = frozenset([link.source, link.target])
        src_intf = _abbreviate_intf(link.source_intf).replace('"', "")
        tgt_intf = _abbreviate_intf(link.target_intf or "").replace('"', "")
        label = f"{src_intf} <-> {tgt_intf}".strip()
        if key not in seen:
            lines.append(f'  "{link.source}" -- "{link.target}" [label="{label}"];')
            seen.add(key)
    lines.append("}")
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# SVG (via Graphviz dot CLI)
# ---------------------------------------------------------------------------


def export_svg(result: TopologyResult) -> bytes:
    """Render topology as SVG using the system graphviz dot binary.

    Raises RuntimeError if dot is not available.
    Raises subprocess.CalledProcessError if rendering fails.
    """
    if shutil.which("dot") is None:
        raise RuntimeError("graphviz dot binary not found — install graphviz to enable SVG export")
    dot_src = export_dot(result)
    proc = subprocess.run(
        ["dot", "-Tsvg"],
        input=dot_src.encode(),
        capture_output=True,
        timeout=30,
    )
    if proc.returncode != 0:
        raise RuntimeError(proc.stderr.decode(errors="replace"))
    return proc.stdout


# ---------------------------------------------------------------------------
# JSON
# ---------------------------------------------------------------------------


def export_json(result: TopologyResult) -> bytes:
    """Serialize the full topology result to JSON bytes."""
    return result.model_dump_json(indent=2).encode()
