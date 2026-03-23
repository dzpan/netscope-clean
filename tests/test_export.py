"""Unit tests for Excel and CSV export — verifying all data categories are included."""

from __future__ import annotations

import csv
import io
import zipfile
from datetime import UTC, datetime

import openpyxl

from backend.export import export_csv_zip, export_excel
from backend.models import (
    ArpEntry,
    Device,
    DeviceStatus,
    InterfaceInfo,
    Link,
    MacTableEntry,
    RouteEntry,
    TopologyResult,
    TrunkInfo,
    VlanInfo,
)


def _make_result() -> TopologyResult:
    """Build a TopologyResult with all data categories populated."""
    return TopologyResult(
        session_id="export-test",
        discovered_at=datetime.now(UTC),
        devices=[
            Device(
                id="SW-01",
                hostname="SW-01",
                mgmt_ip="10.0.0.1",
                platform="C9200L",
                status=DeviceStatus.OK,
                interfaces=[
                    InterfaceInfo(
                        name="GigabitEthernet1/0/1",
                        status="up",
                        vlan="10",
                        speed="1000",
                        duplex="full",
                        description="Uplink",
                        ip_address="10.0.0.1",
                    ),
                    InterfaceInfo(name="GigabitEthernet1/0/2", status="down"),
                ],
                vlans=[
                    VlanInfo(vlan_id="1", name="default", status="active"),
                    VlanInfo(vlan_id="10", name="DATA", status="active"),
                ],
                arp_table=[
                    ArpEntry(
                        ip_address="10.0.0.2",
                        mac_address="aa:bb:cc:dd:ee:ff",
                        interface="Vlan10",
                        entry_type="dynamic",
                    ),
                ],
                mac_table=[
                    MacTableEntry(
                        mac_address="aa:bb:cc:dd:ee:ff",
                        vlan_id="10",
                        interface="GigabitEthernet1/0/1",
                        entry_type="dynamic",
                    ),
                ],
                route_table=[
                    RouteEntry(
                        protocol="C",
                        route_type="connected",
                        destination="10.0.0.0/24",
                        interface="Vlan10",
                    ),
                    RouteEntry(
                        protocol="S",
                        route_type="static",
                        destination="0.0.0.0/0",
                        next_hop="10.0.0.254",
                        interface="Vlan10",
                        metric="1/0",
                    ),
                ],
                trunk_info={
                    "GigabitEthernet1/0/24": TrunkInfo(
                        mode="on",
                        encapsulation="802.1q",
                        status="trunking",
                        native_vlan="1",
                        allowed_vlans="1-4094",
                        active_vlans="1,10",
                    ),
                },
            ),
        ],
        links=[
            Link(
                source="SW-01",
                target="SW-02",
                source_intf="GigabitEthernet1/0/1",
                target_intf="GigabitEthernet0/1",
                protocol="CDP",
            ),
        ],
        failures=[],
    )


# ---------------------------------------------------------------------------
# CSV tests
# ---------------------------------------------------------------------------


class TestCSVExport:
    def test_csv_zip_contains_all_files(self) -> None:
        result = _make_result()
        data = export_csv_zip(result)
        with zipfile.ZipFile(io.BytesIO(data)) as zf:
            names = set(zf.namelist())
        expected = {
            "devices.csv",
            "links.csv",
            "failures.csv",
            "interfaces.csv",
            "vlans.csv",
            "arp_table.csv",
            "mac_table.csv",
            "routes.csv",
            "trunks.csv",
            "etherchannels.csv",
            "stp_vlans.csv",
            "stp_ports.csv",
            "nve_peers.csv",
            "vni_mappings.csv",
            "evpn_neighbors.csv",
        }
        assert expected.issubset(names)

    def _read_csv(self, data: bytes, filename: str) -> list[list[str]]:
        with zipfile.ZipFile(io.BytesIO(data)) as zf:
            content = zf.read(filename).decode()
        reader = csv.reader(io.StringIO(content))
        return list(reader)

    def test_interfaces_csv(self) -> None:
        rows = self._read_csv(export_csv_zip(_make_result()), "interfaces.csv")
        assert rows[0] == [
            "Device",
            "Interface",
            "Status",
            "VLAN",
            "Speed",
            "Duplex",
            "Description",
            "IP Address",
        ]
        assert len(rows) == 3  # header + 2 interfaces
        assert rows[1][1] == "GigabitEthernet1/0/1"

    def test_vlans_csv(self) -> None:
        rows = self._read_csv(export_csv_zip(_make_result()), "vlans.csv")
        assert rows[0] == ["Device", "VLAN ID", "Name", "Status"]
        assert len(rows) == 3  # header + 2 vlans

    def test_arp_table_csv(self) -> None:
        rows = self._read_csv(export_csv_zip(_make_result()), "arp_table.csv")
        assert rows[0] == ["Device", "IP Address", "MAC Address", "Interface", "Type"]
        assert len(rows) == 2  # header + 1 entry
        assert rows[1][1] == "10.0.0.2"

    def test_mac_table_csv(self) -> None:
        rows = self._read_csv(export_csv_zip(_make_result()), "mac_table.csv")
        assert rows[0] == ["Device", "MAC Address", "VLAN", "Interface", "Type"]
        assert len(rows) == 2

    def test_routes_csv(self) -> None:
        rows = self._read_csv(export_csv_zip(_make_result()), "routes.csv")
        assert rows[0] == [
            "Device",
            "Protocol",
            "Route Type",
            "Destination",
            "Next Hop",
            "Interface",
            "Metric",
        ]
        assert len(rows) == 3  # header + 2 routes

    def test_trunks_csv(self) -> None:
        rows = self._read_csv(export_csv_zip(_make_result()), "trunks.csv")
        assert rows[0] == [
            "Device",
            "Port",
            "Mode",
            "Encapsulation",
            "Status",
            "Native VLAN",
            "Allowed VLANs",
            "Active VLANs",
            "Forwarding VLANs",
        ]
        assert len(rows) == 2  # header + 1 trunk
        assert rows[1][1] == "GigabitEthernet1/0/24"


# ---------------------------------------------------------------------------
# Excel tests
# ---------------------------------------------------------------------------


class TestExcelExport:
    def _get_sheet_names(self, data: bytes) -> list[str]:
        wb = openpyxl.load_workbook(io.BytesIO(data))
        return wb.sheetnames

    def _read_sheet(self, data: bytes, sheet_name: str) -> list[list[object]]:
        wb = openpyxl.load_workbook(io.BytesIO(data))
        ws = wb[sheet_name]
        return [[cell.value for cell in row] for row in ws.iter_rows()]

    def test_excel_contains_all_sheets(self) -> None:
        data = export_excel(_make_result())
        names = self._get_sheet_names(data)
        for expected in ["Interfaces", "VLANs", "ARP Table", "MAC Table", "Routes", "Trunks"]:
            assert expected in names, f"Missing sheet: {expected}"

    def test_interfaces_sheet(self) -> None:
        rows = self._read_sheet(export_excel(_make_result()), "Interfaces")
        assert rows[0] == [
            "Device",
            "Interface",
            "Status",
            "VLAN",
            "Speed",
            "Duplex",
            "Description",
            "IP Address",
        ]
        assert len(rows) == 3
        assert rows[1][1] == "GigabitEthernet1/0/1"

    def test_vlans_sheet(self) -> None:
        rows = self._read_sheet(export_excel(_make_result()), "VLANs")
        assert rows[0] == ["Device", "VLAN ID", "Name", "Status"]
        assert len(rows) == 3

    def test_arp_table_sheet(self) -> None:
        rows = self._read_sheet(export_excel(_make_result()), "ARP Table")
        assert rows[0] == ["Device", "IP Address", "MAC Address", "Interface", "Type"]
        assert len(rows) == 2

    def test_mac_table_sheet(self) -> None:
        rows = self._read_sheet(export_excel(_make_result()), "MAC Table")
        assert rows[0] == ["Device", "MAC Address", "VLAN", "Interface", "Type"]
        assert len(rows) == 2

    def test_routes_sheet(self) -> None:
        rows = self._read_sheet(export_excel(_make_result()), "Routes")
        assert rows[0] == [
            "Device",
            "Protocol",
            "Route Type",
            "Destination",
            "Next Hop",
            "Interface",
            "Metric",
        ]
        assert len(rows) == 3

    def test_trunks_sheet(self) -> None:
        rows = self._read_sheet(export_excel(_make_result()), "Trunks")
        assert rows[0][0] == "Device"
        assert rows[0][1] == "Port"
        assert len(rows) == 2
        assert rows[1][1] == "GigabitEthernet1/0/24"

    def test_empty_data_categories(self) -> None:
        """Export with no interfaces/vlans/etc should still create sheets with headers only."""
        result = TopologyResult(
            session_id="empty",
            discovered_at=datetime.now(UTC),
            devices=[
                Device(id="SW-01", hostname="SW-01", mgmt_ip="10.0.0.1", status=DeviceStatus.OK),
            ],
            links=[],
            failures=[],
        )
        data = export_excel(result)
        rows = self._read_sheet(data, "Interfaces")
        assert len(rows) == 1  # header only
